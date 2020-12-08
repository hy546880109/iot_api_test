# !/usr/bin/python3
# -*- coding: utf-8 -*-
# @Author : Huny
# @Email : hy546880109@qq.com
# @date  : 2020.12.08
# @Project: 云平台接口测试用例
from openpyxl import load_workbook

class ParseExcel(object):
    def __init__(self, excelPath, sheetName):
        print(excelPath,sheetName)
        self.wb = load_workbook(excelPath)
        self.sheet = self.wb[sheetName]
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        dataList = []
        for line in self.sheet.rows:
            tmpList=[]
            tmpList.append(line[0].value)
            tmpList.append(line[1].value)
            dataList.append(tmpList)
        return dataList[2:]
